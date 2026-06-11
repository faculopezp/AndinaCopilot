#!/usr/bin/env python3
"""Genera Base_Andina_Ventas_Auto.xlsx desde data/base_nacional.csv.

Uso:
    python scripts/build_excel.py

Hojas:
    Snapshot       — último acumulado por país/marca (una fila por país-marca)
    Peru mensual   — serie mensual Perú (si existe peru_nacional_mensual.csv)
    Chile mensual  — serie mensual Chile (si existe chile_mensual.csv)
    Ecuador mensual— serie mensual Ecuador (si existe ecuador_mensual.csv)
    Prospección H2 — ranking de score SDR para H2 2026
"""
import csv
import json
import pathlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).resolve().parents[1]
DATA = ROOT / "data"

# Colores corporativos
NAVY   = "1F3864"
BLUE   = "2E5496"
LBLUE  = "D9E1F2"
GREEN  = "1E8449"
RED    = "C0392B"
AMBER  = "B9770E"
WHITE  = "FFFFFF"
GREY   = "F4F6FB"

MARCAS_CHINAS = {
    "Changan","Jetour","Geely","Chery","Dfsk","Shineray","Jac","Foton",
    "Great Wall","Haval","Byd","Jmc","Forland","Kyc","King Long","Dongfeng",
    "Gac","Mg","Baic","Jaecoo","Omoda","Sinotruk","Maxus","Dfm","Kaiyi",
    "Leapmotor","Zeekr","Neta","Wuling","Riddara","Deepal",
}


def header_style(cell, bg=NAVY, fg=WHITE, bold=True):
    cell.font = Font(bold=bold, color=fg, size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def autofit(ws, min_w=8, max_w=40):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, min_w), max_w)


def load_csv(filename):
    path = DATA / filename
    if not path.exists():
        return []
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


# ---- Hoja Snapshot ----

def sheet_snapshot(wb):
    rows = load_csv("base_nacional.csv")
    if not rows:
        return
    ws = wb.active
    ws.title = "Snapshot"
    ws.freeze_panes = "A2"

    headers = ["País", "Período", "Marca", "Origen", "Unidades", "Año anterior", "Var% YoY", "Fuente"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(1, i, h)
        header_style(c)

    for r in rows:
        marca = r["marca"]
        origen = "China 🇨🇳" if marca in MARCAS_CHINAS else "Otra"
        try:
            uc = int(r["unidades_curr"])
        except (ValueError, KeyError):
            uc = r.get("unidades_curr") or r.get("uc", "")
        try:
            up = int(r["unidades_prev"]) if r.get("unidades_prev") not in ("", None, "n/d") else None
        except (ValueError, KeyError):
            up = None
        try:
            var = float(r["var_yoy_pct"]) if r.get("var_yoy_pct") not in ("", None, "n/d") else None
        except (ValueError, KeyError):
            var = None

        row_data = [r["pais"], r["periodo"], marca, origen, uc, up if up else "", var if var is not None else "", r["fuente"]]
        ws.append(row_data)
        row_idx = ws.max_row
        # Color var% YoY
        var_cell = ws.cell(row_idx, 7)
        if var is not None:
            if var > 10:
                var_cell.font = Font(color=GREEN, bold=True)
            elif var < -5:
                var_cell.font = Font(color=RED, bold=True)
            var_cell.number_format = '+0.0%;-0.0%;0.0%'
            var_cell.value = var / 100 if var else 0
        # Zebra
        if row_idx % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row_idx, col).fill = PatternFill("solid", fgColor=GREY)

    autofit(ws)


# ---- Hoja serie mensual genérica ----

def sheet_mensual(wb, filename, title):
    rows = load_csv(filename)
    if not rows:
        return
    ws = wb.create_sheet(title)
    ws.freeze_panes = "A2"

    headers = ["País", "Año", "Mes", "Marca", "Origen", "Acum. año", "Mensual"]
    for i, h in enumerate(headers, 1):
        header_style(ws.cell(1, i, h))

    for r in rows:
        marca = r["marca"]
        try:
            acum = int(r["unid_acum"])
        except (ValueError, KeyError):
            acum = ""
        try:
            mensual = int(r["unid_mes"]) if r.get("unid_mes") not in ("", None) else ""
        except (ValueError, KeyError):
            mensual = ""
        ws.append([r["pais"], int(r["anio"]), int(r["mes"]), marca,
                   "China 🇨🇳" if marca in MARCAS_CHINAS else "Otra",
                   acum, mensual])
        row_idx = ws.max_row
        if row_idx % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row_idx, col).fill = PatternFill("solid", fgColor=GREY)

    autofit(ws)


# ---- Hoja Prospección H2 ----

def sheet_prosp(wb):
    rows_snap = load_csv("base_nacional.csv")
    if not rows_snap:
        return

    # Normalizar
    data = []
    for r in rows_snap:
        try:
            uc = int(r.get("unidades_curr") or r.get("uc") or 0)
        except ValueError:
            uc = 0
        try:
            var_raw = r.get("var_yoy_pct") or r.get("var")
            var = float(var_raw) if var_raw not in ("", None, "n/d") else None
        except ValueError:
            var = None
        data.append({"pais": r["pais"], "marca": r["marca"], "uc": uc, "var": var})

    # Vol score por país
    from collections import defaultdict
    por_pais = defaultdict(list)
    for d in data:
        por_pais[d["pais"]].append(d)
    vol_score = {}
    for pais, items in por_pais.items():
        mx = max(x["uc"] for x in items) or 1
        for x in items:
            vol_score[(pais, x["marca"])] = x["uc"] / mx * 100

    # Presencia multi-país
    pais_count = defaultdict(set)
    for d in data:
        pais_count[d["marca"]].add(d["pais"])

    scored = []
    for d in data:
        vs = vol_score.get((d["pais"], d["marca"]), 0)
        var_cap = min(max(d["var"] or 0, 0), 250)
        gs = var_cap / 250 * 100
        mp = (len(pais_count[d["marca"]]) - 1) / 3 * 100
        cn_bonus = 12 if d["marca"] in MARCAS_CHINAS else 0
        score = round(vs * 0.40 + gs * 0.38 + mp * 0.22 + cn_bonus)
        tier = "Alta" if score >= 60 else "Media" if score >= 38 else "Seguimiento"
        scored.append({**d, "score": score, "tier": tier,
                        "npaises": len(pais_count[d["marca"]]),
                        "origen": "China" if d["marca"] in MARCAS_CHINAS else "Otra"})
    scored.sort(key=lambda x: -x["score"])

    ws = wb.create_sheet("Prospeccion H2")
    ws.freeze_panes = "A2"
    headers = ["#", "Marca", "País", "Origen", "Prioridad", "Score", "Unidades", "Var% YoY", "Países"]
    for i, h in enumerate(headers, 1):
        header_style(ws.cell(1, i, h))

    tier_colors = {"Alta": RED, "Media": AMBER, "Seguimiento": "888888"}
    for i, r in enumerate(scored, 1):
        var_str = f"+{r['var']:.1f}%" if r["var"] and r["var"] > 0 else (f"{r['var']:.1f}%" if r["var"] is not None else "n/d")
        ws.append([i, r["marca"], r["pais"], r["origen"], r["tier"],
                   r["score"], r["uc"], var_str, f"{r['npaises']}/4"])
        row_idx = ws.max_row
        # Color tier cell
        tier_cell = ws.cell(row_idx, 5)
        tier_cell.font = Font(color=tier_colors.get(r["tier"], "000000"), bold=True)
        if row_idx % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row_idx, col).fill = PatternFill("solid", fgColor=GREY)

    autofit(ws)


# ---- Entry point ----

def main():
    wb = Workbook()
    sheet_snapshot(wb)
    sheet_mensual(wb, "peru_nacional_mensual.csv", "Peru mensual")
    sheet_mensual(wb, "chile_mensual.csv",          "Chile mensual")
    sheet_mensual(wb, "ecuador_mensual.csv",        "Ecuador mensual")
    sheet_prosp(wb)

    out = ROOT / "Base_Andina_Ventas_Auto.xlsx"
    wb.save(out)
    print(f"OK -> {out}")


if __name__ == "__main__":
    main()
